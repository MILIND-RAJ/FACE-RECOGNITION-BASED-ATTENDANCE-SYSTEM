import face_recognition
import cv2
import numpy
import openpyxl
import time
from datetime import date
from datetime import datetime
import smtplib
from email.message import EmailMessage
import pickle
import os



#yatharth_face = face_recognition.load_image_file("/Users/yathartharora/Downloads/yatharth.jpg")
# yatharth_face_encoding = face_recognition.face_encodings(yatharth_face)[0];
#
# rohan_face  = face_recognition.load_image_file("/Users/yathartharora/Downloads/rohan.jpg")
# rohan_face_encoding = face_recognition.face_encodings(rohan_face)[0];
#
# milind_face  = face_recognition.load_image_file("/Users/yathartharora/Downloads/milind.jpg")
# milind_face_encoding = face_recognition.face_encodings(milind_face)[0];
#
#
# known_face_encodings = [yatharth_face_encoding,rohan_face_encoding,milind_face_encoding]
#
# known_name_face = ["yatharth","rohan","milind"]
#
today = date.today()
d1 = today.strftime("%d")
print(d1)
path = '/Users/yathartharora/January.xlsx'
path_1 = '/Users/yathartharora/Informationfile.xlsx'
live_frame_encodings = []
flag = [0] * 5
matches = []
known_face_encodings=[]
known_name_face=[]

with open("known_face_encodings.pk",'rb') as p:
    known_face_encodings=pickle.load(p)
with open("known_name_face.pk",'rb') as p:
    known_name_face=pickle.load(p)

print(known_name_face)

video_capture = cv2.VideoCapture(0)

while True:
	t = time.localtime()
	s1 = time.strftime("%H",t)
	res,frame = video_capture.read()
	cv2.imshow('Video', frame)
	rgb = frame[: , :, ::-1]
	matches = []
	#live_frame = face_recognition.face_locations(rgb)
	live_frame_encodings = face_recognition.face_encodings(rgb)
	#print(len(live_frame_encodings))
	for i in live_frame_encodings:
		matches = face_recognition.compare_faces(known_face_encodings,i,0.4)
		#print(matches)

		#if True in matches:
		for j in range(len(matches)):
			if matches[j]==True:
				print(known_name_face[j])
				name = known_name_face[j]
				wb = openpyxl.load_workbook(path)
				sheet_obj = wb.active
				#maxrow = sheet_obj.max_row

				for k in range(2,10):
					if(flag[j]==0):
						cell_obj = sheet_obj.cell(row = k,column = 1)
						if(cell_obj.value == name):
							cell_obj1 = sheet_obj.cell(row = k,column = 3*int(d1)-1)
							cell_obj1.value = "Present"
							cell_obj1 = sheet_obj.cell(row = k,column = 3*int(d1))
							t = time.localtime()
							c = time.strftime("%H:%M:%S", t)
							cell_obj1.value = c
							wb.save(path)

							# sending the mail

							wb1 = openpyxl.load_workbook(path_1)
							sheet_obj = wb1.active
							maxrow = sheet_obj.max_row

							# if(flag[i]==0):
							# for j in range(2, 10):
							# 	cell_obj = sheet_obj.cell(row=j,column=1)
							# 	if(cell_obj.value == name):
							cell_obj_1 = sheet_obj.cell(row = k,column = 2)
							s = smtplib.SMTP('smtp.gmail.com', 587)
							s.starttls()
							s.login("testingemail1411@gmail.com","Testmail")
							msg = EmailMessage()
							msg.set_content("You have been marked present")
							msg['Subject'] = "ATTENDENCE CONFIRMED"
							msg['From'] = "testingemail1411@gmail.com"
							msg['To'] = cell_obj_1.value
							s.send_message(msg)
							flag[j] = 1
							s.quit()








							#Out time

				if(int(s1) >= 12):

					for k in range(2,10):
						if(flag[j]==1):
							wb = openpyxl.load_workbook(path)
							sheet_obj = wb.active
							cell_obj = sheet_obj.cell(row = k,column = 1)
							if(cell_obj.value == name):
								cell_obj1 = sheet_obj.cell(row = k,column = 3*int(d1)+1)
								t = time.localtime()
								c = time.strftime("%H:%M:%S", t)
								cell_obj1.value = c
								wb.save(path)

								# sending the mail

								wb1 = openpyxl.load_workbook(path_1)
								sheet_obj = wb1.active
								maxrow = sheet_obj.max_row

								# if(flag[i]==1):
								# for j in range(2, 10):
								# 	cell_obj = sheet_obj.cell(row=j,column=1)
								# 	if(cell_obj.value == name):
								cell_obj_1 = sheet_obj.cell(row = k,column = 2)
								s = smtplib.SMTP('smtp.gmail.com', 587)
								s.starttls()
								s.login("testingemail1411@gmail.com","Testmail")
								msg = EmailMessage()
								msg.set_content("Hope you had a productive day! ")
								msg['Subject'] = "OUT-TIME MARKED"
								msg['From'] = "testingemail1411@gmail.com"
								msg['To'] = cell_obj_1.value
								s.send_message(msg)
								flag[j] = 2
								s.quit()


	if cv2.waitKey(1) & 0xFF == ord('q'):
        	break
video_capture.release()
cv2.destroyAllWindows()
